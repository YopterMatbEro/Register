import openpyxl, os, re, logging, datetime as dt


def get_names_in_dir(path):
    return [filename for filename in os.listdir(path)]


class MyExcelTable(openpyxl.Workbook):
    __new_doc__ = """
        obj.get_requests_dict() - возвращает библиотеку со всеми заявками в указанной директории
        obj.save_to_doc() - вносит данные на лист и сохраняет документ
                """

    def help(self):
        print(self.__new_doc__)

    def __init__(self, name, path):
        super().__init__()
        self.name = name
        self.path = path
        self.workbook = openpyxl.load_workbook(name)
        self.STATIONS = ['Роза Хутор', 'Эсто-Садок', 'Имеретинский курорт', 'Олимпийская деревня',
                         'Адлер', 'Аэропорт', 'Хоста', 'Мацеста', 'Сочи']

    def get_requests_dict(self):
        absolute_path = self.path
        all_requests = {}
        dir_names = get_names_in_dir(absolute_path)
        for directory in dir_names:
            path = absolute_path + '/' + directory
            file_names = get_names_in_dir(path)
            for filename in file_names:
                pattern = r'^(\d+)\s+акт\s+(\d+[.]?\d*)\s+(.+)\s+расхождение по\s+.+\s+(\d+[,\d+]*)\s+руб\. от (\d+[.,]\d+[.,]\d+)'
                match = re.match(pattern, filename)
                if match:
                    request_number = match.group(1)
                    act_number = match.group(2)
                    station_name = match.group(3)
                    amount = match.group(4)
                    date = match.group(5)
                    if request_number in all_requests.keys():
                        print(f'Номер заявки: {request_number} уже значится в библиотеке. Проверьте данные:')
                        print(f'Номер Акта: {act_number}', end='; ')
                        print(f'Станция: {station_name}', end='; ')
                        print(f'Сумма: {amount}', end='; ')
                        print(f'Дата: {date}')
                    else:
                        all_requests[request_number] = [4, act_number, station_name, date, directory, amount]
                else:
                    pattern = r'^(\d+)\s+акт\s+(\d+[\.]?\d+)\s+(\w+[ -]?[\w+]*)\s+[нулевая|отсутвутет].+(\d{2}[.,]\d{2}[.,]\d{4})'
                    match = re.match(pattern, filename)
                    if match:
                        request_number = match.group(1)
                        act_number = match.group(2)
                        station_name = match.group(3)
                        if ' ' in station_name and station_name.split()[1] == 'нулевая':
                            station_name = station_name.split()[0]
                        date = match.group(4)
                        if request_number in all_requests.keys():
                            print(f'Номер заявки: {request_number} уже значится в библиотеке. Проверьте данные:')
                            print(f'Номер Акта: {act_number}', end='; ')
                            print(f'Станция: {station_name}', end='; ')
                            print(f'Дата: {date}')
                        else:
                            all_requests[request_number] = [3, act_number, station_name, date, directory]
        return all_requests

    def save_to_doc(self):
        if os.path.exists('logfile.log'):
            os.remove('logfile.log')
        logging.basicConfig(filename='logfile.log', level=logging.INFO,
                            format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        all_requests = self.get_requests_dict()
        sheet = self.workbook.get_sheet_by_name('май 2023 ')
        for row in sheet.iter_rows(min_row=6):
            if row[0].value in self.STATIONS:
                temp = row[0].value
            if row[5].value is not None and type(row[5].value) == int:
                for key, value in all_requests.items():
                    if value[2] == temp and value[1] == row[5].value:
                        row[11].value = key
                        row[12].value = value[4]
                        break  # Exit the inner loop after finding a match
                else:
                    logging.info(f'row[0].value = {row[0].value} - no matching request found')
            else:
                logging.info(f'row[0].value = {row[0].value} - empty row')

        self.workbook.save(self.name)

        # for row in sheet.iter_rows(min_row=6):
        #     row_values = [cell.value for cell in row]
        #     for i, value in enumerate(row_values):
        #         sheet.cell(row=row[0].row, column=i + 1, value=value)
